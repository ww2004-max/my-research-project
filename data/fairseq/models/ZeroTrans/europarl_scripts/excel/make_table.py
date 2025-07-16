import openpyxl
import os

root_path="../europarl_scripts"
language_sequence = ["en", "de", "fi", "pt", "bg", 'sl', "it",  "pl", "hu", "ro", "es", "da", "nl", "et", "cs"]
language_dict = {'en': 1, 'de': 2, 'nl': 3, 'da': 4, 'es': 5, 'pt': 6, 'ro': 7,
                 'it': 8, 'sl': 9, 'pl': 10, 'cs': 11, 'bg': 12, "fi": 13, "hu": 14, "et": 15}



def _read_txt_strip_(url):
    file = open(url, 'r', encoding='utf-8')
    lines = file.readlines()
    file.close()
    return [line.strip() for line in lines]


# type = scores or off
def mk_margin(start_row, type, sheet):
    sheet.cell(row=start_row, column=2).value = 'tgt'
    sheet.cell(row=(start_row + 1), column=2).value = type
    sheet.cell(row=(start_row + 1), column=1).value = 'src'
    for idx, lang in enumerate(language_sequence):
        sheet.cell(row=(start_row + 2 + idx), column=2).value = lang
        sheet.cell(row=(start_row + 1), column=(3 + idx)).value = lang


# data_1 is list(list * 15)
# inner list len is len(language_sequence) - 1
# data_2 is list[en2m, m2en, sup, zero]
def fill_data(start_row, data_1, sheet):
    for i in range(len(language_sequence)):
        tmp_list = data_1[i]
        for j in range(len(language_sequence)):
            if i == j: continue
            idx = j if i > j else (j - 1)
            sheet.cell(row=(start_row + 2 + i), column=(3 + j)).value = tmp_list[idx]


# final data is [list * 4] corresponding to bleu, p, r, f
def make_bar(sheet, final_data):
    sheet.cell(row=1, column=20).value = "en2m"
    sheet.cell(row=1, column=21).value = "m2en"
    sheet.cell(row=1, column=22).value = "supervised"
    sheet.cell(row=1, column=23).value = "zero"
    sequence = ["bleu", "p", "r", "f"]
    for idx, tmp in enumerate(final_data):
        sheet.cell(row=(idx + 2), column=19).value = sequence[idx]
        sheet.cell(row=(idx + 2), column=20).value = tmp[0]
        sheet.cell(row=(idx + 2), column=21).value = tmp[1]
        sheet.cell(row=(idx + 2), column=22).value = tmp[2]
        sheet.cell(row=(idx + 2), column=23).value = tmp[3]


def extract_results_from_file(model_id, dir_name):
    dir_path = os.path.join(root_path, "results", dir_name)
    final_results_list =[0, 0, 0, 0]
    data = _read_txt_strip_(os.path.join(root_path, dir_path, str(model_id), "{}.sacrebleu".format(model_id)))

    bleu_list = [[0 for _ in range(len(language_sequence) - 1)] for _ in range(len(language_sequence))]
    # en2m, m2en, supervised, zero
    idx_i, idx_j = '', ''
    for row in data:
        if "-" in row and len(row) < 10:
            print(row)
            tmp = row.split("-")
            idx_i, idx_j = language_dict[tmp[0]] - 1, language_dict[tmp[1]] - 1
        if "\"score\"" in row:
            score = float(row.split(":")[1].rstrip(","))
            score = round(score, 2)
            if idx_i > idx_j:
                bleu_list[idx_i][idx_j] = score
            else:
                bleu_list[idx_i][idx_j - 1] = score

            if idx_i == 0:
                final_results_list[0] += score
            elif idx_j == 0:
                final_results_list[1] += score
            else:
                final_results_list[3] += score
    final_results_list[0] = round(final_results_list[0]/(len(language_sequence) - 1), 2)
    final_results_list[1] = round(final_results_list[1] / (len(language_sequence) - 1), 2)
    final_results_list[2] = round(((final_results_list[0] + final_results_list[1]) / 2), 2)
    final_results_list[3] = round(final_results_list[3] / ((len(language_sequence) - 1) * (len(language_sequence) - 2)), 2)

    data = _read_txt_strip_(os.path.join(root_path, dir_path, str(model_id), "{}.bertscore".format(model_id)))
    p_list = [[0 for _ in range(len(language_sequence) - 1)] for _ in range(len(language_sequence))]
    r_list = [[0 for _ in range(len(language_sequence) - 1)] for _ in range(len(language_sequence))]
    f_list = [[0 for _ in range(len(language_sequence) - 1)] for _ in range(len(language_sequence))]
    idx_i, idx_j = '', ''
    for row in data:
        if "-" in row:
            print(row)
            tmp = row.split("-")
            idx_i, idx_j = language_dict[tmp[0]] - 1, language_dict[tmp[1]] - 1
        if "P:" in row:
            tmp = row.split(" ")
            p, r, f = float(tmp[1].strip()), float(tmp[3].strip()), float(tmp[5].strip())
            if idx_i > idx_j:
                p_list[idx_i][idx_j], r_list[idx_i][idx_j], f_list[idx_i][idx_j] = p, r, f
            else:
                p_list[idx_i][idx_j-1], r_list[idx_i][idx_j-1], f_list[idx_i][idx_j-1] = p, r, f
            if idx_i == 0:
                final_results_list[1][0] += p
                final_results_list[2][0] += r
                final_results_list[3][0] += f
            elif idx_j == 0:
                final_results_list[1][1] += p
                final_results_list[2][1] += r
                final_results_list[3][1] += f
            else:
                final_results_list[1][3] += p
                final_results_list[2][3] += r
                final_results_list[3][3] += f
    final_results_list[1][0] = round(final_results_list[1][0] / (len(language_sequence) - 1), 2)
    final_results_list[1][1] = round(final_results_list[1][1] / (len(language_sequence) - 1), 2)
    final_results_list[1][2] = round(((final_results_list[1][0] + final_results_list[1][1]) / 2), 2)
    final_results_list[1][3] = round(final_results_list[1][3] / ((len(language_sequence) - 1) * (len(language_sequence) - 2)), 2)

    final_results_list[2][0] = round(final_results_list[2][0] / (len(language_sequence) - 1), 2)
    final_results_list[2][1] = round(final_results_list[2][1] / (len(language_sequence) - 1), 2)
    final_results_list[2][2] = round(((final_results_list[2][0] + final_results_list[2][1]) / 2), 2)
    final_results_list[2][3] = round(final_results_list[2][3] / ((len(language_sequence) - 1) * (len(language_sequence) - 2)), 2)

    final_results_list[3][0] = round(final_results_list[3][0] / (len(language_sequence) - 1), 2)
    final_results_list[3][1] = round(final_results_list[3][1] / (len(language_sequence) - 1), 2)
    final_results_list[3][2] = round(((final_results_list[3][0] + final_results_list[3][1]) / 2), 2)
    final_results_list[3][3] = round(final_results_list[3][3] / ((len(language_sequence) - 1) * (len(language_sequence) - 2)), 2)

    return bleu_list, p_list, r_list, f_list, final_results_list


language_family = {'germanic': [2, 3, 4],
                   'romance': [5, 6, 7, 8],
                   'slavic': [9, 10, 11, 12],
                   'uralic': [13, 14, 15]}


def make_bar_inner_f(sheet, inner_families):
    sequence = ["bleu", "p", "r", "f"]
    families_list = list(language_family.keys())
    start_row = 19
    sheet.cell(row=start_row, column=19).value = "inner"
    for i in range(len(families_list)):
        sheet.cell(row=start_row, column=(20 + i)).value = families_list[i]

    # 4 metrics
    for i in range(len(inner_families)):
        tmp = inner_families[i]
        sheet.cell(row=(start_row + 1 + i), column=19).value = sequence[i]
        sheet.cell(row=(start_row + 1 + i), column=20).value = round(tmp[0], 2)
        sheet.cell(row=(start_row + 1 + i), column=21).value = round(tmp[1], 2)
        sheet.cell(row=(start_row + 1 + i), column=22).value = round(tmp[2], 2)
        sheet.cell(row=(start_row + 1 + i), column=23).value = round(tmp[3], 2)


def make_bar_f2f(sheet, families2families):
    sequence = ["bleu", "p", "r", "f"]
    families_list = list(language_family.keys())
    start_row = 25
    for i in range(len(families_list)):
        results = families2families[i]
        sheet.cell(row=(start_row + i * 6), column=19).value = families_list[i]
        residual = [lang for lang in families_list if lang != families_list[i]]
        for k in range(3):
            sheet.cell(row=(start_row + i * 6), column=(20 + k)).value = residual[k]
        # 4 metrics
        for j in range(4):
            tmp = results[j]
            sheet.cell(row=(start_row + i * 6 + 1 + j), column=19).value = sequence[j]
            sheet.cell(row=(start_row + i * 6 + 1 + j), column=20).value = round(tmp[0], 2)
            sheet.cell(row=(start_row + i * 6 + 1 + j), column=21).value = round(tmp[1], 2)
            sheet.cell(row=(start_row + i * 6 + 1 + j), column=22).value = round(tmp[2], 2)


def make_bar_en_families(sheet, en2families, families2en):
    sequence = ["bleu", "p", "r", "f"]
    sheet.cell(row=7, column=20).value = "germanic"
    sheet.cell(row=7, column=21).value = "romance"
    sheet.cell(row=7, column=22).value = "slavic"
    sheet.cell(row=7, column=23).value = "uralic"
    sheet.cell(row=7, column=19).value = "en2f"
    for idx, tmp in enumerate(en2families):
        sheet.cell(row=(idx + 8), column=19).value = sequence[idx]
        sheet.cell(row=(idx + 8), column=20).value = round(tmp[0], 2)
        sheet.cell(row=(idx + 8), column=21).value = round(tmp[1], 2)
        sheet.cell(row=(idx + 8), column=22).value = round(tmp[2], 2)
        sheet.cell(row=(idx + 8), column=23).value = round(tmp[3], 2)

    sheet.cell(row=13, column=20).value = "germanic"
    sheet.cell(row=13, column=21).value = "romance"
    sheet.cell(row=13, column=22).value = "slavic"
    sheet.cell(row=13, column=23).value = "uralic"
    sheet.cell(row=13, column=19).value = "f2en"
    for idx, tmp in enumerate(families2en):
        sheet.cell(row=(idx + 14), column=19).value = sequence[idx]
        sheet.cell(row=(idx + 14), column=20).value = round(tmp[0], 2)
        sheet.cell(row=(idx + 14), column=21).value = round(tmp[1], 2)
        sheet.cell(row=(idx + 14), column=22).value = round(tmp[2], 2)
        sheet.cell(row=(idx + 14), column=23).value = round(tmp[3], 2)


def mk_table(method_name, model_id):
    bleu_list, p_list, r_list, f_list, final_results = extract_results_from_file(method_name, model_id)
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(index=0, title="sheet1")
    mk_margin(1, "bleu", sheet)
    fill_data(1, bleu_list, sheet)

    mk_margin(19, "bs-p", sheet)
    fill_data(19, p_list, sheet)

    mk_margin(38, "bs-r", sheet)
    fill_data(38, r_list, sheet)

    mk_margin(57, "bs-f", sheet)
    fill_data(57, f_list, sheet)

    make_bar(sheet, final_results)

    # [[4 families] * 4 metrics]
    en2families, families2en = [[0, 0, 0, 0] for _ in range(4)], [[0, 0, 0, 0] for _ in range(4)]
    # # 4 metrics: [ 4 (src) * 3 (tgt) = 12] * 4
    families_list = list(language_family.keys())
    for i in range(len(families_list)):
        languages_of_this_family = language_family[families_list[i]]
        for j in languages_of_this_family:
            en2families[0][i] += bleu_list[0][j - 2] / len(languages_of_this_family)
            en2families[1][i] += p_list[0][j - 2] / len(languages_of_this_family)
            en2families[2][i] += r_list[0][j - 2] / len(languages_of_this_family)
            en2families[3][i] += f_list[0][j - 2] / len(languages_of_this_family)

            families2en[0][i] += bleu_list[j - 1][0] / len(languages_of_this_family)
            families2en[1][i] += p_list[j - 1][0] / len(languages_of_this_family)
            families2en[2][i] += r_list[j - 1][0] / len(languages_of_this_family)
            families2en[3][i] += f_list[j - 1][0] / len(languages_of_this_family)
    make_bar_en_families(sheet, en2families, families2en)

    # [[4 families] * 4 metrics]
    inner_families = [[0, 0, 0, 0] for _ in range(4)]
    for i in range(len(families_list)):
        languages_of_this_family = language_family[families_list[i]]
        for j in languages_of_this_family:
            for k in languages_of_this_family:
                if j == k: continue
                denominator = len(languages_of_this_family) * (len(languages_of_this_family) - 1)
                idx_src = j - 1
                idx_tgt = k - 1 if k < j else k - 2
                inner_families[0][i] += bleu_list[idx_src][idx_tgt] / denominator
                inner_families[1][i] += p_list[idx_src][idx_tgt] / denominator
                inner_families[2][i] += r_list[idx_src][idx_tgt] / denominator
                inner_families[3][i] += f_list[idx_src][idx_tgt] / denominator

    make_bar_inner_f(sheet, inner_families)

    families2families = [[[0, 0, 0] for _ in range(4)] for _ in range(4)]
    for i in range(len(families_list)):
        languages_of_src_family = language_family[families_list[i]]
        for j in range(len(families_list)):
            if i == j: continue
            # idx_filling, 若j>i, 则j-1
            idx_filling = j if j < i else j - 1
            languages_of_tgt_family = language_family[families_list[j]]
            for src_lang in languages_of_src_family:
                for tgt_lang in languages_of_tgt_family:
                    idx_src = src_lang - 1
                    idx_tgt = tgt_lang - 1 if tgt_lang < src_lang else tgt_lang - 2
                    denominator = len(languages_of_tgt_family) * len(languages_of_src_family)
                    families2families[i][0][idx_filling] += bleu_list[idx_src][idx_tgt] / denominator
                    families2families[i][1][idx_filling] += p_list[idx_src][idx_tgt] / denominator
                    families2families[i][2][idx_filling] += r_list[idx_src][idx_tgt] / denominator
                    families2families[i][3][idx_filling] += f_list[idx_src][idx_tgt] / denominator
    make_bar_f2f(sheet, families2families)
    os.mkdir(os.path.join(root_path, "excel", method_name))
    wb.save(os.path.join(root_path, "excel", method_name, f"{model_id}.xlsx"))

dir_name = "vanilla"
start_id = 1
end_id = 1
for i in range(1, end_id + 1, 1):
    mk_table(dir_name, i)