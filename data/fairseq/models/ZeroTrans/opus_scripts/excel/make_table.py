import openpyxl
import os

root_path = ".."
zero_languages = ["de", "nl", "ar", "fr", "zh", "ru",]
zero_dict = {"de":0, "nl":1, "ar":2, "fr":3, "zh":4, "ru":5,}
language_dict = {"en": 1,"af": 2,"am": 3,"ar": 4,"as": 5,"az": 6,"be": 7,"bg": 8,"bn": 9,"br": 10,
                         "bs": 11,"ca": 12,"cs": 13,"cy": 14,"da": 15,"de": 16,"el": 17,"eo": 18,"es": 19,"et": 20,
                         "eu": 21,"fa": 22,"fi": 23,"fr": 24,"fy": 25,"ga": 26,"gd": 27,"gl": 28,"gu": 29,"ha": 30,
                         "he": 31,"hi": 32,"hr": 33,"hu": 34,"id": 35,"ig": 36,"is": 37,"it": 38,"ja": 39,"ka": 40,
                         "kk": 41,"km": 42,"kn": 43,"ko": 44,"ku": 45,"ky": 46,"li": 47,"lt": 48,"lv": 49,"mg": 50,
                         "mk": 51,"ml": 52,"mr": 53,"ms": 54,"mt": 55,"my": 56,"nb": 57,"ne": 58,"nl": 59,"nn": 60,
                         "no": 61,"oc": 62,"or": 63,"pa": 64,"pl": 65,"ps": 66,"pt": 67,"ro": 68,"ru": 69,"rw": 70,
                         "se": 71,"sh": 72,"si": 73,"sk": 74,"sl": 75,"sq": 76,"sr": 77,"sv": 78,"ta": 79,"te": 80,
                         "tg": 81,"th": 82,"tk": 83,"tr": 84,"tt": 85,"ug": 86,"uk": 87,"ur": 88,"uz": 89,"vi": 90,
                         "wa": 91,"xh": 92,"yi": 93,"zh": 94,"zu": 95}


def _read_txt_strip_(url):
    file = open(url, 'r', encoding='utf-8')
    lines = file.readlines()
    file.close()
    return [line.strip() for line in lines]


def extract_results_from_file(method_name, model_id):
    dir_path = f"{root_path}/opus_scripts/results"
    final_results_list = [0, 0, 0, 0]
    data = _read_txt_strip_(os.path.join( dir_path, str(method_name), str(model_id), "{}.sacrebleu".format(model_id)))
    # en-> & ->en
    supervised_bleu_list = [[0 for _ in range(len(language_dict.keys()) - 1)] for _ in range(2)]
    zero_bleu_list = [[0 for _ in range(len(zero_languages))] for _ in range(len(zero_languages))]
    idx_i, idx_j = None, None
    zero_flag = False
    for row in data:
        if "-" in row and len(row) < 10:
            tmp = row.split("-")
            if tmp[0] == "en" and tmp[1] != "en":
                idx_i, idx_j = 0, language_dict[tmp[1]] - 2
                zero_flag = False
            elif tmp[0] != "en" and tmp[1] == "en":
                idx_i, idx_j = 1, language_dict[tmp[0]] - 2
                zero_flag = False
            else:
                idx_i, idx_j = zero_dict[tmp[0]], zero_dict[tmp[1]]
                zero_flag = True
        if "\"score\"" in row:
            score = float(row.split(":")[1].rstrip(","))
            score = round(score, 2)
            if not zero_flag:
                supervised_bleu_list[idx_i][idx_j] = score
            else:
                zero_bleu_list[idx_i][idx_j] = score

            if idx_i == 0:
                final_results_list[0] += score
            elif idx_i == 1:
                final_results_list[1] += score
            else:
                final_results_list[3] += score
    final_results_list[0] = round(final_results_list[0] / (len(language_dict.keys()) - 1), 2)
    final_results_list[1] = round(final_results_list[1] / (len(language_dict.keys()) - 1), 2)
    final_results_list[2] = round(((final_results_list[0] + final_results_list[1]) / 2), 2)
    final_results_list[3] = round(final_results_list[3] / (len(zero_languages) * (len(zero_languages)-1)), 2)

    data = _read_txt_strip_(os.path.join(dir_path, str(method_name), str(model_id), "{}.bertscore".format(model_id)))
    p_list = [[0 for _ in range(len(zero_languages))] for _ in range(len(zero_languages))]
    r_list = [[0 for _ in range(len(zero_languages))] for _ in range(len(zero_languages))]
    f_list = [[0 for _ in range(len(zero_languages))] for _ in range(len(zero_languages))]
    bert_score_list = [0, 0, 0]
    idx_i, idx_j = None, None
    sum_p, sum_r, sum_f = 0,0,0
    for row in data:
        if "-" in row:
            tmp = row.split("-")
            idx_i, idx_j = zero_dict[tmp[0]], zero_dict[tmp[1]]
        if "P:" in row:
            tmp = row.split(" ")
            p, r, f = float(tmp[1].strip()), float(tmp[3].strip()), float(tmp[5].strip())
            p_list[idx_i][idx_j], r_list[idx_i][idx_j], f_list[idx_i][idx_j] = p, r, f
            sum_p += p
            sum_r += r
            sum_f += f
    bert_score_list[0], bert_score_list[1], bert_score_list[2] = round(sum_p/(len(zero_languages) * (len(zero_languages)-1)), 2),\
                                                                 round(sum_r/(len(zero_languages) * (len(zero_languages)-1)), 2),\
                                                                 round(sum_f/(len(zero_languages) * (len(zero_languages)-1)), 2)
    return supervised_bleu_list, zero_bleu_list, p_list, r_list, f_list, final_results_list, bert_score_list


def mk_table(method_name, model_id):
    supervised_bleu_list, zero_bleu_list, p_list, r_list, f_list, final_results_list, bert_score_list = extract_results_from_file(method_name, model_id)
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(index=0, title="sheet1")
    sheet.cell(row=1, column=1).value = 'supervised'
    sheet.cell(row=1, column=2).value = 'sacrebleu:'
    sheet.cell(row=1, column=3).value, sheet.cell(row=1, column=4).value, sheet.cell(row=1, column=5).value = "en->x",\
                                                                                                              "x->en",\
                                                                                                              "sum",
    sheet.cell(row=2, column=3).value, sheet.cell(row=2, column=4).value, sheet.cell(row=2, column=5).value = final_results_list[0],\
                                                                                                              final_results_list[1],\
                                                                                                              final_results_list[2]
    sheet.cell(row=3, column=1).value = 'en->x'
    sheet.cell(row=6, column=1).value = 'x->en'
    for idx, lang in enumerate(list(language_dict.keys())):
        if lang == "en": continue
        sheet.cell(row=4, column=(1 + idx)).value = lang
        sheet.cell(row=5, column=(1 + idx)).value = supervised_bleu_list[0][idx - 1]
        sheet.cell(row=7, column=(1 + idx)).value = lang
        sheet.cell(row=8, column=(1 + idx)).value = supervised_bleu_list[1][idx - 1]

    sheet.cell(row=10, column=1).value = 'zero-shot'
    sheet.cell(row=10, column=2).value = 'sacrebleu'
    sheet.cell(row=10, column=3).value = 'p'
    sheet.cell(row=10, column=4).value = 'r'
    sheet.cell(row=10, column=5).value = 'f'
    sheet.cell(row=11, column=2).value = final_results_list[3]
    sheet.cell(row=11, column=3).value, sheet.cell(row=11, column=4).value, sheet.cell(row=11, column=5).value = bert_score_list[0],\
                                                                                                                 bert_score_list[1],\
                                                                                                                 bert_score_list[2]

    sheet.cell(row=12, column=1).value = 'sacrebleu'
    for i in range(len(zero_languages)):
        sheet.cell(row=(14 + i), column=1).value = zero_languages[i]
        sheet.cell(row=13, column=(2 + i)).value = zero_languages[i]
        for j in range(len(zero_languages)):
            if zero_bleu_list[i][j] != 0:
                sheet.cell(row=(14 + i), column=(2 + j)).value = zero_bleu_list[i][j]

    sheet.cell(row=20, column=1).value = 'p'
    for i in range(len(zero_languages)):
        sheet.cell(row=(22 + i), column=1).value = zero_languages[i]
        sheet.cell(row=21, column=(2 + i)).value = zero_languages[i]
        for j in range(len(zero_languages)):
            if zero_bleu_list[i][j] != 0:
                sheet.cell(row=(22 + i), column=(2 + j)).value = p_list[i][j]

    sheet.cell(row=28, column=1).value = 'r'
    for i in range(len(zero_languages)):
        sheet.cell(row=(30 + i), column=1).value = zero_languages[i]
        sheet.cell(row=29, column=(2 + i)).value = zero_languages[i]
        for j in range(len(zero_languages)):
            if zero_bleu_list[i][j] != 0:
                sheet.cell(row=(30 + i), column=(2 + j)).value = r_list[i][j]

    sheet.cell(row=36, column=1).value = 'f'
    for i in range(len(zero_languages)):
        sheet.cell(row=(38 + i), column=1).value = zero_languages[i]
        sheet.cell(row=37, column=(2 + i)).value = zero_languages[i]
        for j in range(len(zero_languages)):
            if zero_bleu_list[i][j] != 0:
                sheet.cell(row=(38 + i), column=(2 + j)).value = f_list[i][j]
    wb.save("{}/opus_scripts/excel/{}/".format(root_path, method_name) + str(model_id) + "_opus.xlsx")


dir_name = "zero"
start_id = 1
end_id = 1
for i in range(1, end_id + 1, 1):
    mk_table(dir_name, i)