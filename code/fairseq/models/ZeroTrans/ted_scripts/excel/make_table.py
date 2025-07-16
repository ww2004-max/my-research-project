import openpyxl
import os

language_sequence = ["en", "ar", "he", "ru", "ko", "it", "ja", "zh", "es", "nl", "vi", "tr", "fr", "pl", "ro", "fa", "hr", "cs", "de"]
language_dict = {'en': 1, 'ar': 2, 'he': 3, 'ru': 4, 'ko': 5, 'it': 6, 'ja': 7,
                         'zh': 8, 'es': 9, 'nl': 10, "vi": 11, "tr": 12, "fr":13,
                         "pl": 14, "ro":15, "fa":16, "hr":17, "cs":18, "de": 19}

root_path = ".."

def _read_txt_strip_(url):
    file = open(url, 'r', encoding='utf-8')
    lines = file.readlines()
    file.close()
    return [line.strip() for line in lines]


# type = scores or off
def mk_margin(start_row, type, sheet):
    sheet.cell(row=start_row, column=2).value = 'tgt'
    sheet.cell(row=(start_row + 1), column=2).value = type
    sheet.cell(row=(start_row + 1), column=1).value = 'src'
    for idx, lang in enumerate(language_sequence):
        sheet.cell(row=(start_row + 2 + idx), column=2).value = lang
        sheet.cell(row=(start_row + 1), column=(3 + idx)).value = lang


# data_1 is list(list * 15)
# inner list len is len(language_sequence) - 1
# data_2 is list[en2m, m2en, sup, zero]
def fill_data(start_row, data_1, sheet):
    for i in range(len(language_sequence)):
        tmp_list = data_1[i]
        for j in range(len(language_sequence)):
            if i == j: continue
            idx = j if i > j else (j - 1)
            sheet.cell(row=(start_row + 2 + i), column=(3 + j)).value = tmp_list[idx]

def make_bar(sheet, final_data):
    lang_len = len(language_sequence)
    sheet.cell(row=1, column=(lang_len+5)).value = "en2m"
    sheet.cell(row=1, column=(lang_len+6)).value = "m2en"
    sheet.cell(row=1, column=(lang_len+7)).value = "supervised"
    sheet.cell(row=1, column=(lang_len+8)).value = "zero"
    sequence = ["bleu", "p", "r", "f"]
    for idx, tmp in enumerate(final_data):
        sheet.cell(row=(idx + 2), column=(lang_len+4)).value = sequence[idx]
        sheet.cell(row=(idx + 2), column=(lang_len+5)).value = tmp[0]
        sheet.cell(row=(idx + 2), column=(lang_len+6)).value = tmp[1]
        sheet.cell(row=(idx + 2), column=(lang_len+7)).value = tmp[2]
        sheet.cell(row=(idx + 2), column=(lang_len+8)).value = tmp[3]

def extract_results_from_file(model_id, dir_name):
    dir_path = os.path.join(root_path, "ted_scripts", "results", dir_name)
    final_results_list = [[0, 0, 0, 0] for _ in range(4)]
    data = _read_txt_strip_(os.path.join(dir_path, str(model_id), "{}.sacrebleu".format(model_id)))

    bleu_list = [[0 for _ in range(len(language_sequence) - 1)] for _ in range(len(language_sequence))]
    # en2m, m2en, supervised, zero
    idx_i, idx_j = '', ''
    for row in data:
        if "-" in row and len(row) < 10:
            print(row)
            tmp = row.split("-")
            idx_i, idx_j = language_dict[tmp[0]] - 1, language_dict[tmp[1]] - 1
        if "\"score\"" in row:
            if len(row) < 30:
                score = float(row.split(":")[1].rstrip(","))
                score = round(score, 2)
            else:
                score = float(row.split(":")[2].split(",")[0])
                score = round(score, 2)
            if idx_i > idx_j:
                bleu_list[idx_i][idx_j] = score
            else:
                bleu_list[idx_i][idx_j - 1] = score

            if idx_i == 0:
                final_results_list[0][0] += score
            elif idx_j == 0:
                final_results_list[0][1] += score
            else:
                final_results_list[0][3] += score
    final_results_list[0][0] = round(final_results_list[0][0]/(len(language_sequence) - 1), 2)
    final_results_list[0][1] = round(final_results_list[0][1] / (len(language_sequence) - 1), 2)
    final_results_list[0][2] = round(((final_results_list[0][0] + final_results_list[0][1]) / 2), 2)
    final_results_list[0][3] = round(final_results_list[0][3] / ((len(language_sequence) - 1) * (len(language_sequence) - 2)), 2)

    data = _read_txt_strip_(os.path.join(dir_path, str(model_id), "{}.bertscore".format(model_id)))
    p_list = [[0 for _ in range(len(language_sequence) - 1)] for _ in range(len(language_sequence))]
    r_list = [[0 for _ in range(len(language_sequence) - 1)] for _ in range(len(language_sequence))]
    f_list = [[0 for _ in range(len(language_sequence) - 1)] for _ in range(len(language_sequence))]
    idx_i, idx_j = '', ''
    for row in data:
        if "-" in row:
            print(row)
            tmp = row.split("-")
            idx_i, idx_j = language_dict[tmp[0]] - 1, language_dict[tmp[1]] - 1
        if "P:" in row:
            tmp = row.split(" ")
            p, r, f = float(tmp[1].strip()), float(tmp[3].strip()), float(tmp[5].strip())
            if idx_i > idx_j:
                p_list[idx_i][idx_j], r_list[idx_i][idx_j], f_list[idx_i][idx_j] = p, r, f
            else:
                p_list[idx_i][idx_j-1], r_list[idx_i][idx_j-1], f_list[idx_i][idx_j-1] = p, r, f
            if idx_i == 0:
                final_results_list[1][0] += p
                final_results_list[2][0] += r
                final_results_list[3][0] += f
            elif idx_j == 0:
                final_results_list[1][1] += p
                final_results_list[2][1] += r
                final_results_list[3][1] += f
            else:
                final_results_list[1][3] += p
                final_results_list[2][3] += r
                final_results_list[3][3] += f
    final_results_list[1][0] = round(final_results_list[1][0] / (len(language_sequence) - 1), 2)
    final_results_list[1][1] = round(final_results_list[1][1] / (len(language_sequence) - 1), 2)
    final_results_list[1][2] = round(((final_results_list[1][0] + final_results_list[1][1]) / 2), 2)
    final_results_list[1][3] = round(final_results_list[1][3] / ((len(language_sequence) - 1) * (len(language_sequence) - 2)), 2)

    final_results_list[2][0] = round(final_results_list[2][0] / (len(language_sequence) - 1), 2)
    final_results_list[2][1] = round(final_results_list[2][1] / (len(language_sequence) - 1), 2)
    final_results_list[2][2] = round(((final_results_list[2][0] + final_results_list[2][1]) / 2), 2)
    final_results_list[2][3] = round(final_results_list[2][3] / ((len(language_sequence) - 1) * (len(language_sequence) - 2)), 2)

    final_results_list[3][0] = round(final_results_list[3][0] / (len(language_sequence) - 1), 2)
    final_results_list[3][1] = round(final_results_list[3][1] / (len(language_sequence) - 1), 2)
    final_results_list[3][2] = round(((final_results_list[3][0] + final_results_list[3][1]) / 2), 2)
    final_results_list[3][3] = round(final_results_list[3][3] / ((len(language_sequence) - 1) * (len(language_sequence) - 2)), 2)

    return bleu_list, p_list, r_list, f_list, final_results_list


def mk_table(model_id, dir_name):
    bleu_list, p_list, r_list, f_list, final_results = extract_results_from_file(model_id, dir_name)
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(index=0, title="sheet1")
    mk_margin(1, "bleu", sheet)
    fill_data(1, bleu_list, sheet)

    mk_margin(23, "bs-p", sheet)
    fill_data(23, p_list, sheet)

    mk_margin(45, "bs-r", sheet)
    fill_data(45, r_list, sheet)

    mk_margin(67, "bs-f", sheet)
    fill_data(67, f_list, sheet)

    make_bar(sheet, final_results)

    save_dir = os.path.join(root_path, "ted_scripts", "excel", dir_name)
    print(save_dir)
    if not os.path.exists(save_dir):
        os.makedirs(save_dir)
    
    wb.save(os.path.join(save_dir, str(model_id) + ".xlsx"))

dir_name = "zero"
start_id = 1
end_id = 1
for i in range(1, end_id + 1, 1):
    mk_table(i, dir_name)